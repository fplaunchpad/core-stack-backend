import requests
from nrm_app.settings import GEOSERVER_URL, EXCEL_PATH
from utilities.gee_utils import valid_gee_text
import xml.etree.ElementTree as ET
from lxml import etree as LET
from nrm_app.celery import app
from computing.models import *
from utilities.geoserver_utils import Geoserver
import json
from django.conf import settings
from pathlib import Path
from utilities.constants import (
    GEOSERVER_BASE,
)
from utilities.logger import setup_logger
from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from rest_framework.response import Response
from rest_framework import status
from .utils import (
    _is_cache_valid,
    _set_cache,
    send_report_email,
)
import os
import openpyxl
from stats_generator.models import LayerInfo

logger = setup_logger(__name__)


def get_url(geoserver_url, workspace, layer_name):
    return (
        f"{geoserver_url}/{workspace}/ows"
        f"?service=WFS"
        f"&version=1.1.0"
        f"&request=GetFeature"
        f"&typeName={workspace}:{layer_name}"
        f"&resultType=hits"
    )


def load_workspace_config():
    """
    Load workspace configuration from JSON file.
    """
    config_path = (
        Path(settings.BASE_DIR)
        / "data"
        / "layers"
        / "layer_status"
        / "layer_mapping.json"
    )
    with open(config_path, "r") as f:
        return json.load(f)


@app.task(bind=True)
def layer_status(self, state, district, block):
    """
    Check the status of all layers for a particular location.

    Args:
        self: Instance reference
        state: State name
        district: District name
        block: Block name

    Returns:
        Dictionary with status of all workspace layers
    """
    print(f"{state=}")
    all_workspace_statuses = {}
    capabilities_cache = {}
    district = valid_gee_text(district.lower())
    block = valid_gee_text(block.lower())

    # Load workspace configuration from JSON
    workspace_config = load_workspace_config()

    for workspace_display, config in workspace_config.items():
        workspace = config.get("name")
        suffix = config.get("suffix", "")
        prefix = config.get("prefix", "")
        layer_type = config.get("type", "")

        # constructing layer name
        layer_name_parts = [prefix, district, block, suffix]
        layer_name = "_".join(part for part in layer_name_parts if part)

        total_features = 0
        end_date = None
        start_date = None
        status_code = 400
        # checking for vector layer
        if layer_type == "vector":
            layer_url = get_url(GEOSERVER_URL, workspace, layer_name)
            res_layer_url = requests.get(layer_url)

            if res_layer_url.status_code == 200:
                try:
                    root = ET.fromstring(res_layer_url.text)
                    # Extract feature count from WFS hits response
                    total_features = int(root.attrib.get("numberOfFeatures", 0))
                    status_code = 200 if total_features > 0 else 400
                    layer = (
                        Layer.objects.filter(layer_name=layer_name)
                        .order_by("-layer_version")
                        .first()
                    )
                    if layer and layer.misc:
                        start_date = layer.misc.get("start_date")
                        end_date = layer.misc.get("end_date")

                except ET.ParseError:
                    print(f"Invalid XML for layer: {layer_name}")
                    status_code = 400
        else:
            if workspace not in capabilities_cache:
                capabilities_url = f"{GEOSERVER_BASE}{workspace}/wms?service=WMS&request=GetCapabilities"
                try:
                    response = requests.get(capabilities_url, timeout=30)
                    if response.status_code == 200:
                        parser = LET.XMLParser(recover=True, encoding="utf-8")
                        root = LET.fromstring(response.content, parser=parser)
                        ns = {"wms": root.tag.split("}")[0].strip("{")}
                        layers = root.findall(".//wms:Layer/wms:Name", namespaces=ns)
                        capabilities_cache[workspace] = {
                            layer.text for layer in layers if layer.text
                        }
                    else:
                        capabilities_cache[workspace] = set()
                except Exception as e:
                    print(
                        f"Failed to fetch capabilities for workspace {workspace}: {e}"
                    )
                    capabilities_cache[workspace] = set()

            if layer_name in capabilities_cache.get(workspace, set()):
                status_code = 200
        all_workspace_statuses[workspace_display] = {
            "workspace": workspace,
            "layer_name": layer_name,
            "status_code": status_code,
            "totalFeature": total_features,
            "endDate": end_date,
            "startDate": start_date,
        }

    return all_workspace_statuses


def load_workspace_types():
    """
    Load workspace types configuration from JSON file.
    """
    config_path = (
        Path(settings.BASE_DIR)
        / "data"
        / "layers"
        / "workspace_layers"
        / "layers_in_workspace.json"
    )
    with open(config_path, "r") as f:
        return json.load(f)


@app.task(bind=True)
def get_layers_of_workspace(self, workspace):
    """
    It will take workspace as argument and returns all the layers which is present on geoserver.
    """
    # Load workspace types from JSON
    workspace_types = load_workspace_types()
    raster_workspace = workspace_types["raster_workspace"]
    vector_workspace = workspace_types["vector_workspace"]
    raster_and_vector_workspace = workspace_types["raster_and_vector_workspace"]

    geo = Geoserver()
    layers = geo.get_layers(workspace)
    layer_names = [layer["name"] for layer in layers["layers"]["layer"]]
    if workspace in raster_workspace:
        print("you passed raster workspace")
        available_layers = valid_raster_layers_for_workspace(workspace)
        valid_layers = [ln for ln in layer_names if ln in available_layers]
        invalid_layers = [ln for ln in layer_names if ln not in available_layers]
        return {"valid_layer": valid_layers, "invalid_layers": invalid_layers}
    elif workspace in vector_workspace:
        print("you passed vector workspace")
        valid_layers = []
        invalid_layers = []
        for layer_name in layer_names:
            if is_valid_vector_layer(workspace, layer_name):
                valid_layers.append(layer_name)
            else:
                invalid_layers.append(layer_name)
        return {"valid_layer": valid_layers, "invalid_layers": invalid_layers}
    elif workspace in raster_and_vector_workspace:
        print("you passed workspace which contain both layers(raster and vector)")
        valid_layers = []
        invalid_layers = []
        for layer_name in layer_names:
            if "vector" in layer_name.lower():
                if is_valid_vector_layer(workspace, layer_name):
                    valid_layers.append(layer_name)
                else:
                    invalid_layers.append(layer_name)
            elif "raster" in layer_name.lower():
                available_layers = valid_raster_layers_for_workspace(workspace)
                if layer_name in available_layers:
                    valid_layers.append(layer_name)
                else:
                    invalid_layers.append(layer_name)
        return {"valid_layer": valid_layers, "invalid_layers": invalid_layers}
    else:
        print("you passed wrong workspace")
        return []


_raster_cache = {}
_vector_cache = {}


def valid_raster_layers_for_workspace(workspace: str) -> set:
    if _is_cache_valid(_raster_cache, workspace):
        logger.info(f"Cache hit for raster: {workspace}")
        return _raster_cache[workspace]["data"]

    session = get_session_with_retry()
    capabilities_url = (
        f"{GEOSERVER_BASE}{workspace}/wms?service=WMS&request=GetCapabilities"
    )
    try:
        response = session.get(capabilities_url, timeout=(5, 15))
        response.raise_for_status()
    except requests.exceptions.Timeout:
        logger.warning(
            f"Timeout fetching raster capabilities for {workspace} — not caching"
        )
        return set()
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed raster capabilities for {workspace}: {e} — not caching")
        return set()

    root = ET.fromstring(response.content)
    ns = {"wms": root.tag.split("}")[0].strip("{")}
    layers = root.findall(".//wms:Layer/wms:Name", namespaces=ns)
    result = {layer.text for layer in layers}

    _set_cache(_raster_cache, workspace, result)  # cache with timestamp
    logger.info(f"Cached raster layers for {workspace}: {len(result)} layers")
    return result


def valid_vector_layers_for_workspace(workspace: str) -> set:
    if _is_cache_valid(_vector_cache, workspace):
        logger.info(f"Cache hit for vector: {workspace}")
        return _vector_cache[workspace]["data"]

    session = get_session_with_retry()
    capabilities_url = (
        f"{GEOSERVER_BASE}{workspace}/wfs"
        f"?service=WFS&version=2.0.0&request=GetCapabilities"
    )
    try:
        response = session.get(capabilities_url, timeout=(5, 15))
        response.raise_for_status()
    except requests.exceptions.Timeout:
        logger.warning(
            f"Timeout fetching vector capabilities for {workspace} — not caching"
        )
        return set()
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed vector capabilities for {workspace}: {e} — not caching")
        return set()

    root = ET.fromstring(response.content)
    ns = {"wfs": "http://www.opengis.net/wfs/2.0"}
    names = root.findall(".//wfs:FeatureType/wfs:Name", namespaces=ns)
    result = {name.text.split(":")[-1] for name in names}

    _set_cache(_vector_cache, workspace, result)  # cache with timestamp
    logger.info(f"Cached vector layers for {workspace}: {len(result)} layers")
    return result


def is_valid_vector_layer(workspace: str, layer_name: str) -> bool:
    """Used in parallel fallback only."""
    session = get_session_with_retry()
    try:
        layer_url = get_url(GEOSERVER_URL, workspace, layer_name)
        res = session.get(layer_url, timeout=(5, 10))  # shorter timeout per layer
        if res.status_code == 200:
            root = ET.fromstring(res.text)
            return int(root.attrib.get("numberOfFeatures", 0)) > 0
    except requests.exceptions.Timeout:
        logger.warning(f"Timeout checking vector layer: {layer_name}")
    except Exception as e:
        logger.warning(f"Vector check failed for {layer_name}: {e}")
    return False


def bulk_check_vector_layers(
    workspace: str,
    layer_names: list[str],
    max_workers: int = 20,
) -> dict[str, bool]:
    """
    Checks multiple vector layers concurrently using a thread pool.
    Returns {layer_name: is_valid} mapping.
    """
    results = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_layer = {
            executor.submit(is_valid_vector_layer, workspace, name): name
            for name in layer_names
        }
        for future in as_completed(future_to_layer):
            layer_name = future_to_layer[future]
            try:
                results[layer_name] = future.result()
            except Exception as e:
                logger.warning(f"Failed check for {layer_name}: {e}")
                results[layer_name] = False
    return results


@app.task(bind=True)
def missing_layer_for_all_workspace(self):
    dataset_qs = Dataset.objects.filter(workspace__isnull=False, is_active=True)
    workspaces = dataset_qs.values_list("workspace", flat=True).distinct()
    can_be_empty_workspaces = (
        dataset_qs.filter(can_be_empty=True)
        .values_list("workspace", flat=True)
        .distinct()
    )
    workspaces = [w.strip() for w in workspaces if w and w.strip()]
    can_be_empty = set(w.strip() for w in can_be_empty_workspaces if w and w.strip())
    result = {"Mandatory": {}, "can_be_empty": {}}
    for workspace in workspaces:
        layer_result = check_missing_layers(workspace)
        if workspace in can_be_empty:
            result["can_be_empty"][workspace] = layer_result
        else:
            result["Mandatory"][workspace] = layer_result
    send_report_email(result, report_type="missing_layers")
    return result


def check_missing_layers(workspace: str) -> dict:
    logger.info(f"{workspace=}")
    workspace_config = load_workspace_config()
    workspace_types = get_workspace_types(workspace)
    logger.info(f"Found types: {workspace_types}")

    if not workspace_types:
        logger.critical(f"No config found for workspace: {workspace}")
        return {"no config found": []}

    layer_config = get_layer_config_by_type(
        workspace_config, workspace, workspace_types
    )

    # ── Fetch all available layers ONCE (bulk, cached) ──────────────────────
    available_raster_layers = valid_raster_layers_for_workspace(workspace)
    available_vector_layers = valid_vector_layers_for_workspace(workspace)  # bulk WFS`
    use_bulk_vector = bool(available_vector_layers)  # fallback if WFS unavailable

    # ── Pre-build all (tehsil, layer_type, layer_name, state) combos ────────
    active_tehsils = TehsilSOI.objects.select_related(
        "district__state"  # eliminates N+1 DB queries
    ).filter(active_status=True)

    tasks = []  # (state, layer_type, layer_name)
    for tehsil_obj in active_tehsils:
        state = tehsil_obj.district.state.state_name
        district_name = valid_gee_text(tehsil_obj.district.district_name.lower())
        tehsil_name = valid_gee_text(tehsil_obj.tehsil_name.lower())

        for layer_type, configs in layer_config.items():
            for config in configs:
                prefix = config.get("prefix")
                suffix = config.get("suffix")
                layer_name = "_".join(
                    p for p in [prefix, district_name, tehsil_name, suffix] if p
                )
                tasks.append(
                    (state, district_name, tehsil_name, layer_type, layer_name)
                )

    # ── Check raster layers (pure set lookup, no HTTP) ───────────────────────
    missing_layers = []
    vector_tasks = []  # collect for batch/parallel processing

    for state, district_name, tehsil_name, layer_type, layer_name in tasks:
        if layer_type == "raster":
            if layer_name not in available_raster_layers:
                missing_layers.append(
                    f"{state}, {district_name}, {tehsil_name}, {layer_name}"
                )

        elif layer_type == "vector":
            if use_bulk_vector:
                # O(1) set lookup — no HTTP call needed
                if layer_name not in available_vector_layers:
                    missing_layers.append(
                        f"{state}, {district_name}, {tehsil_name}, {layer_name}"
                    )
            else:
                vector_tasks.append(
                    (state, district_name, tehsil_name, layer_name)
                )  # queue for parallel check

    # ── Parallel fallback for vector layers (if WFS bulk fetch failed) ───────
    if vector_tasks:
        layer_names = [ln for _, _, _, ln in vector_tasks]
        info_by_name = {ln: (st, dn, tn) for st, dn, tn, ln in vector_tasks}
        validity = bulk_check_vector_layers(workspace, layer_names)

        for layer_name, is_valid in validity.items():
            if not is_valid:
                state, district_name, tehsil_name = info_by_name[layer_name]
                missing_layers.append(
                    f"{state}, {district_name}, {tehsil_name}, {layer_name}"
                )

    return {"missing_layers": missing_layers}


def get_workspace_types(workspace_name):
    """
    Return all layer types for a given workspace name from DB.
    Example: ['raster', 'vector']
    """
    return list(
        Dataset.objects.filter(workspace=workspace_name)
        .values_list("layer_type", flat=True)
        .distinct()
    )


def get_layer_config_by_type(workspace_config, workspace_name, layer_types):
    """
    Return list of prefix/suffix configs for each layer type.

    Args:
        workspace_config (dict): config JSON
        workspace_name (str): dataset name
        layer_types (list): ['raster', 'vector']

    Returns:
        dict: {'raster': [{'prefix': ..., 'suffix': ...}, ...], 'vector': [...]}
    """
    result = {}

    for config in workspace_config.values():
        if config.get("name") == workspace_name and config.get("type") in layer_types:
            layer_type = config["type"]

            if layer_type not in result:
                result[layer_type] = []

            result[layer_type].append(
                {
                    "prefix": config.get("prefix"),
                    "suffix": config.get("suffix"),
                }
            )

    return result


def get_session_with_retry():
    """Creates a requests session with retry and timeout handling."""
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1,  # waits 1s, 2s, 4s between retries
        status_forcelist=[500, 502, 503, 504],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def clear_layer_cache(workspace: str = None):
    if workspace:
        _raster_cache.pop(workspace, None)
        _vector_cache.pop(workspace, None)
        logger.info(f"Cleared cache for {workspace}")
    else:
        _raster_cache.clear()
        _vector_cache.clear()
        logger.info("Cleared all layer caches")


def refresh_layer_cache(request, workspace=None):
    clear_layer_cache(workspace)
    return Response({"message": f"Cache cleared for: {workspace or 'all workspaces'}"})


def check_xlsx_sheets(file_path):
    """
    Returns dict with:
    - missing_sheets: expected sheets not present in file
    - empty_sheets: sheets that exist but have no data rows
    """
    missing_sheets = []
    empty_sheets = []
    conditional_sheets = []
    dataset_qs = LayerInfo.objects.filter(
        workspace__isnull=False, excel_to_be_generated=True
    )
    mandatory_sheet_name = (
        dataset_qs.filter(can_be_absent=False)
        .values_list("sheet_name", flat=True)
        .distinct()
    )
    conditional_sheet_name = (
        dataset_qs.filter(can_be_absent=True)
        .values_list("sheet_name", flat=True)
        .distinct()
    )
    EXPECTED_SHEETS = list(
        dict.fromkeys(
            s.strip() for item in mandatory_sheet_name for s in item.split(",")
        )
    )
    CONDITIONAL_DATA_SHEETS = list(
        dict.fromkeys(
            s.strip() for item in conditional_sheet_name for s in item.split(",")
        )
    )
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        existing_sheets = wb.sheetnames

        for sheet_name in EXPECTED_SHEETS + CONDITIONAL_DATA_SHEETS:
            if sheet_name not in existing_sheets:
                if sheet_name in CONDITIONAL_DATA_SHEETS:
                    conditional_sheets.append(sheet_name)
                else:
                    missing_sheets.append(sheet_name)
            else:
                ws = wb[sheet_name]
                # max_row is None or 1 (header only) means no data
                if ws.max_row is None or ws.max_row <= 1:
                    empty_sheets.append(sheet_name)

        # Flag unexpected sheets that are empty too
        for sheet_name in existing_sheets:
            if sheet_name not in EXPECTED_SHEETS:
                ws = wb[sheet_name]
                if ws.max_row is None or ws.max_row <= 1:
                    empty_sheets.append(f"{sheet_name} (unexpected + empty)")

        wb.close()

    except Exception as e:
        return {"error": str(e)}

    return {
        "missing_sheets": missing_sheets,
        "empty_sheets": empty_sheets,
        "conditional_sheets": conditional_sheets,
    }


@app.task(bind=True)
def check_missing_excel_files(self):
    """
    Check missing excel and json files for active tehsils.
    Groups all missing files per tehsil location.
    Also checks xlsx for missing/empty sheets.
    """
    logger.info("inside check missing excel for active locations")

    base_path = os.path.join(EXCEL_PATH, "data/stats_excel_files")
    missing_location = []

    active_tehsils = TehsilSOI.objects.filter(active_status=True).select_related(
        "district__state"
    )

    for tehsil in active_tehsils:
        state = valid_gee_text(tehsil.district.state.state_name.lower())
        district = valid_gee_text(tehsil.district.district_name.lower())
        tehsil_name = valid_gee_text(tehsil.tehsil_name.lower())

        dir_path = os.path.join(base_path, state.upper(), district.upper())

        required_files = [
            f"{district}_{tehsil_name}.json",
            f"{district}_{tehsil_name}.xlsx",
            f"{district}_{tehsil_name}_KYL_filter_data.json",
            f"{district}_{tehsil_name}_KYL_village_data.json",
        ]

        missing_files = []
        xlsx_issues = {}

        for filename in required_files:
            file_path = os.path.join(dir_path, filename)

            if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
                missing_files.append(filename)

            # Deep check for xlsx if file exists and is non-empty
            elif filename.endswith(".xlsx"):
                sheet_check = check_xlsx_sheets(file_path)
                if sheet_check.get("error"):
                    missing_files.append(
                        f"{filename} (unreadable: {sheet_check['error']})"
                    )
                elif (
                    sheet_check["missing_sheets"]
                    or sheet_check["empty_sheets"]
                    or sheet_check["conditional_sheets"]
                ):
                    xlsx_issues = {
                        "file": filename,
                        "missing_sheets": sheet_check["missing_sheets"],
                        "empty_sheets": sheet_check["empty_sheets"],
                        "conditional_sheets": sheet_check["conditional_sheets"],
                    }

        if missing_files or xlsx_issues:
            missing_location.append(
                {
                    "state": state,
                    "district": district,
                    "tehsil": tehsil_name,
                    "missing_files": missing_files,
                    "xlsx_issues": xlsx_issues,
                }
            )
    send_report_email(missing_location, report_type="missing_excel_files")
    logger.info("report sent")
    return missing_location
